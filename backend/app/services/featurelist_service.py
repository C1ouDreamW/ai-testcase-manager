"""FeatureList（功能清单）导出与导入，支持 Markdown 与 Excel 两种格式。

- Markdown：表格形式，多行字段用 <br> 编码换行，单元格内 | 用 \\| 转义，纯文本、Git 友好。
- Excel(xlsx)：所见即所得表格，多行字段自动换行，测试同学在 Excel 里维护更直观。

导出按 format 选择；导入按文件扩展名自动识别。
"""

from io import BytesIO

from app.services.document_parser import DocumentParseError

# 列顺序即导出/导入约定
COLUMNS = [
    ("module", "模块"),
    ("feature", "功能点"),
    ("priority", "优先级"),
    ("description", "功能描述"),
    ("acceptance_criteria", "验收标准"),
    ("constraints", "约束/边界"),
]
HEADER_TO_FIELD = {label: field for field, label in COLUMNS}
COL_WIDTHS = [16, 22, 10, 40, 40, 30]

MEDIA_MD = "text/markdown; charset=utf-8"
MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------- Markdown

def _encode_md_cell(value: str) -> str:
    """将单元格文本编码为 Markdown 表格安全格式，换行转为 `<br>`，竖线与反斜杠转义。

    Args:
        value (str): 原始单元格文本。

    Returns:
        str: 编码后的 Markdown 安全文本。
    """
    text = (value or "").strip()
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\r\n", "\n").replace("\n", "<br>")


def _decode_md_cell(value: str) -> str:
    """将 Markdown 表格单元格解码为原始文本，还原 `<br>`、竖线和反斜杠。

    Args:
        value (str): Markdown 编码后的单元格文本。

    Returns:
        str: 解码后的原始文本。
    """
    text = (value or "").strip()
    text = text.replace("<br>", "\n").replace("<br/>", "\n")
    text = text.replace("\\|", "|").replace("\\\\", "\\")
    return text.strip()


def _split_md_row(line: str) -> list[str]:
    """解析 Markdown 表格行，按竖线分割并处理转义，返回单元格列表。

    Args:
        line (str): Markdown 表格行字符串。

    Returns:
        list[str]: 去除首尾空格的单元格列表。
    """
    cells: list[str] = []
    buf = ""
    escaped = False
    for ch in line:
        if escaped:
            buf += ch
            escaped = False
            continue
        if ch == "\\":
            buf += ch
            escaped = True
            continue
        if ch == "|":
            cells.append(buf)
            buf = ""
            continue
        buf += ch
    cells.append(buf)
    if cells and cells[0].strip() == "":
        cells = cells[1:]
    if cells and cells[-1].strip() == "":
        cells = cells[:-1]
    return cells


def _is_separator_row(cells: list[str]) -> bool:
    """判断 Markdown 表格行是否为分隔行（如 `|---|---|`）。

    Args:
        cells (list[str]): 已分割的单元格列表。

    Returns:
        bool: 是否为分隔行。
    """
    return bool(cells) and all(set(c.strip()) <= {"-", ":", " "} and "-" in c for c in cells)


def export_featurelist_md(title: str, items: list[dict]) -> str:
    """将功能清单导出为 Markdown 表格格式的字符串。

    Args:
        title (str): 文档标题。
        items (list[dict]): 功能点字典列表。

    Returns:
        str: Markdown 表格文本。
    """
    labels = [label for _, label in COLUMNS]
    lines = [
        f"# {title or '功能清单'}",
        "",
        "| " + " | ".join(labels) + " |",
        "| " + " | ".join(["---"] * len(labels)) + " |",
    ]
    for item in items:
        row = [_encode_md_cell(item.get(field, "")) for field, _ in COLUMNS]
        lines.append("| " + " | ".join(row) + " |")
    lines.append("")
    return "\n".join(lines)


def parse_featurelist_md(data: bytes) -> list[dict]:
    """从 Markdown 表格中解析功能清单。

    自动识别编码（UTF-8-BOM、UTF-8、GBK），按表头列名匹配字段。

    Args:
        data (bytes): Markdown 文件的原始字节数据。

    Returns:
        list[dict]: 功能点字典列表。

    Raises:
        DocumentParseError: 编码无法识别、缺少表格或缺少「功能点」列时抛出。
    """
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise DocumentParseError("无法识别文件编码，请使用 UTF-8")

    table_lines = [ln for ln in text.splitlines() if ln.strip().startswith("|")]
    if not table_lines:
        raise DocumentParseError("未找到 Markdown 表格，请使用「导出清单」得到的格式")

    header_cells = [c.strip() for c in _split_md_row(table_lines[0])]
    col_index = {HEADER_TO_FIELD[name]: i for i, name in enumerate(header_cells) if name in HEADER_TO_FIELD}
    if "feature" not in col_index:
        raise DocumentParseError("FeatureList 缺少「功能点」列")

    items: list[dict] = []
    for line in table_lines[1:]:
        cells = _split_md_row(line)
        if _is_separator_row([c.strip() for c in cells]):
            continue

        def cell(field: str) -> str:
            idx = col_index.get(field)
            if idx is None or idx >= len(cells):
                return ""
            return _decode_md_cell(cells[idx])

        feature = cell("feature")
        if not feature:
            continue
        items.append(_build_item(cell))

    if not items:
        raise DocumentParseError("FeatureList 中未解析到任何功能点")
    return items


# ---------------------------------------------------------------- Excel

def export_featurelist_xlsx(title: str, items: list[dict]) -> bytes:
    """将功能清单导出为 .xlsx 格式的字节数据。

    Args:
        title (str): 工作表名称。
        items (list[dict]): 功能点字典列表。

    Returns:
        bytes: .xlsx 文件的字节数据。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "FeatureList"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS[idx - 1]
    ws.freeze_panes = "A2"

    for r, item in enumerate(items, start=2):
        for c, (field, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=item.get(field, "") or "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_featurelist_xlsx(data: bytes) -> list[dict]:
    """从 .xlsx 文件中解析功能清单，按表头列名匹配字段。

    Args:
        data (bytes): .xlsx 文件的原始字节数据。

    Returns:
        list[dict]: 功能点字典列表。

    Raises:
        DocumentParseError: 文件格式错误、为空或缺少「功能点」列时抛出。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentParseError("FeatureList 解析失败，请确认是 .xlsx 文件") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise DocumentParseError("FeatureList 为空")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_index = {HEADER_TO_FIELD[name]: i for i, name in enumerate(header) if name in HEADER_TO_FIELD}
    if "feature" not in col_index:
        raise DocumentParseError("FeatureList 缺少「功能点」列")

    items: list[dict] = []
    for row in rows[1:]:
        def cell(field: str) -> str:
            idx = col_index.get(field)
            if idx is None or idx >= len(row):
                return ""
            value = row[idx]
            return str(value).strip() if value is not None else ""

        feature = cell("feature")
        if not feature:
            continue
        items.append(_build_item(cell))

    if not items:
        raise DocumentParseError("FeatureList 中未解析到任何功能点")
    return items


# ---------------------------------------------------------------- shared / dispatch

def _build_item(cell) -> dict:
    """用 cell 取值函数构建标准化的功能点字典，自动校验优先级合法性。

    Args:
        cell (callable): 以字段名为参数的取值函数。

    Returns:
        dict: 包含 module、feature、description、acceptance_criteria、constraints、priority 的字典。
    """
    priority = cell("priority") or "P1"
    if priority not in ("P0", "P1", "P2"):
        priority = "P1"
    return {
        "module": cell("module"),
        "feature": cell("feature"),
        "description": cell("description"),
        "acceptance_criteria": cell("acceptance_criteria"),
        "constraints": cell("constraints"),
        "priority": priority,
    }


def export_featurelist(title: str, items: list[dict], fmt: str = "xlsx") -> tuple[bytes, str, str]:
    """导出功能清单的统一入口，根据格式返回字节内容、媒体类型和文件扩展名。

    Args:
        title (str): 文档标题。
        items (list[dict]): 功能点字典列表。
        fmt (str, optional): 导出格式，支持 "xlsx" 和 "md"。默认为 "xlsx"。

    Returns:
        tuple[bytes, str, str]: (字节内容, MIME 类型, 文件扩展名)。
    """
    if fmt == "md":
        text = export_featurelist_md(title, items)
        return text.encode("utf-8"), MEDIA_MD, "md"
    content = export_featurelist_xlsx(title, items)
    return content, MEDIA_XLSX, "xlsx"


def parse_featurelist(filename: str, data: bytes) -> list[dict]:
    """根据文件扩展名自动分发到对应的解析器，解析功能清单。

    Args:
        filename (str): 文件名，用于判断格式。
        data (bytes): 文件原始字节数据。

    Returns:
        list[dict]: 功能点字典列表。

    Raises:
        DocumentParseError: 不支持的格式时抛出。
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return parse_featurelist_xlsx(data)
    if name.endswith(".md") or name.endswith(".markdown"):
        return parse_featurelist_md(data)
    raise DocumentParseError("仅支持 .xlsx / .md / .markdown 格式的 FeatureList")
