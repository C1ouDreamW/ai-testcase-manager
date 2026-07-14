"""通用测试用例导出服务，支持 Markdown 与 Excel 两种格式。

- 输入：统一的用例 dict 列表（生成草稿 GeneratedCaseDraft 或已入库 TestCase 均可通过 build_export_row 转换）
- 输出：字节内容 + media_type + 扩展名，可直接作为 HTTP 响应
- 列顺序：编号 / 模块 / 功能点 / 标题 / 优先级 / 类型 / 冒烟 / 前置条件 / 操作步骤 / 预期结果 [/ 评审状态 / 来源]

评审状态列仅在导出「单次生成结果」时启用（include_review=True）。
"""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

BASE_COLUMNS = [
    ("no", "编号"),
    ("module", "模块"),
    ("feature", "功能点"),
    ("title", "标题"),
    ("priority", "优先级"),
    ("case_type", "类型"),
    ("is_smoke", "冒烟"),
    ("precondition", "前置条件"),
    ("steps_text", "操作步骤"),
    ("expected_result", "预期结果"),
]
REVIEW_COLUMNS = [
    ("review_status", "评审状态"),
    ("source", "来源"),
]

COL_WIDTHS = {
    "no": 8,
    "module": 14,
    "feature": 20,
    "title": 36,
    "priority": 8,
    "case_type": 8,
    "is_smoke": 8,
    "precondition": 24,
    "steps_text": 40,
    "expected_result": 32,
    "review_status": 10,
    "source": 12,
}

CASE_TYPE_LABEL = {"functional": "功能", "boundary": "边界", "exception": "异常"}
REVIEW_LABEL = {
    "pending": "待评审",
    "adopted": "已采纳",
    "rejected": "已驳回",
    "edited": "已编辑",
}
SOURCE_LABEL = {
    "ai_generated": "AI 生成",
    "manual": "手动新增",
    "imported": "导入",
}

MEDIA_MD = "text/markdown; charset=utf-8"
MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------- row builder


def _steps_to_text(raw: Any) -> str:
    """将 list 或 JSON 字符串或纯文本统一格式化为带编号的多行步骤文本。

    Args:
        raw (Any): 原始步骤数据，支持 list、JSON 字符串、纯文本。

    Returns:
        str: "1. xxx\n2. xxx" 格式的步骤文本。
    """
    if raw is None:
        return ""
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, str):
        text = raw.strip()
        if not text:
            return ""
        try:
            parsed = json.loads(text)
            items = parsed if isinstance(parsed, list) else [text]
        except json.JSONDecodeError:
            items = [line.strip() for line in text.splitlines() if line.strip()]
    else:
        items = [str(raw)]
    items = [str(s).strip() for s in items if str(s).strip()]
    return "\n".join(f"{i + 1}. {s}" for i, s in enumerate(items))


def build_export_row(case: Any, index: int) -> dict:
    """将 GeneratedCaseDraft、TestCase 或前端 dict 归一化为导出行字典。

    支持 dict 和 SQLAlchemy 对象（用 getattr 兜底）。
    module 和 feature 需上层调用者提前补入 dict 或 ORM 对象。

    Args:
        case (Any): 用例对象，支持 dict 或 SQLAlchemy ORM 对象。
        index (int): 用例序号（从 0 开始）。

    Returns:
        dict: 包含 no、module、feature、title、priority、steps_text 等字段的导出行。
    """

    def _get(key: str, default: Any = "") -> Any:
        if isinstance(case, dict):
            return case.get(key, default)
        return getattr(case, key, default)

    is_smoke = bool(_get("is_smoke", False))
    case_type = _get("case_type", "functional") or "functional"
    review_status = _get("review_status", "") or ""
    source = _get("source", "") or ""

    return {
        "no": index + 1,
        "module": (_get("module", "") or "").strip() or "未分类",
        "feature": (_get("feature", "") or "").strip() or "未指定",
        "title": (_get("title", "") or "").strip(),
        "priority": _get("priority", "P2") or "P2",
        "case_type": CASE_TYPE_LABEL.get(case_type, case_type),
        "is_smoke": "是" if is_smoke else "否",
        "precondition": (_get("precondition", "") or "").strip(),
        "steps_text": _steps_to_text(_get("steps", "")),
        "expected_result": (_get("expected_result", "") or "").strip(),
        "review_status": REVIEW_LABEL.get(review_status, review_status),
        "source": SOURCE_LABEL.get(source, source),
    }


def _columns(include_review: bool) -> list[tuple[str, str]]:
    """获取导出列定义，根据是否包含评审信息决定列集合。

    Args:
        include_review (bool): 是否包含评审状态和来源列。

    Returns:
        list[tuple[str, str]]: 列定义列表，每项为 (字段名, 列标题)。
    """
    return BASE_COLUMNS + (REVIEW_COLUMNS if include_review else [])


# ---------------------------------------------------------------- Markdown (大纲式)


def _strip_step_number(step: str) -> str:
    """去掉步骤前的编号前缀（如"1. xxx"），便于用 Markdown 列表重新展示。

    Args:
        step (str): 步骤文本。

    Returns:
        str: 去除编号前缀后的步骤文本。
    """
    if ". " in step:
        head, tail = step.split(". ", 1)
        if head.isdigit():
            return tail
    return step


def _group_rows(
    rows: list[dict],
) -> tuple[list[str], dict[str, list[str]], dict[str, dict[str, list[dict]]]]:
    """按模块和功能点二级分组，保留首次出现的顺序。

    Args:
        rows (list[dict]): 导出行的列表。

    Returns:
        tuple: (module_order, feature_order, grouped) 三元组。
    """
    module_order: list[str] = []
    feature_order: dict[str, list[str]] = {}
    grouped: dict[str, dict[str, list[dict]]] = {}
    for row in rows:
        mod = row.get("module") or "未分类"
        feat = row.get("feature") or "未指定"
        if mod not in grouped:
            grouped[mod] = {}
            feature_order[mod] = []
            module_order.append(mod)
        if feat not in grouped[mod]:
            grouped[mod][feat] = []
            feature_order[mod].append(feat)
        grouped[mod][feat].append(row)
    return module_order, feature_order, grouped


def _split_text_lines(text: str) -> list[str]:
    """将多行文本按换行拆分为非空行列表。

    Args:
        text (str): 原始文本。

    Returns:
        list[str]: 去除空白后的非空行列表。
    """
    return [line.strip() for line in (text or "").splitlines() if line.strip()]


def export_testcases_md(
    title: str, rows: list[dict], include_review: bool = False
) -> str:
    """大纲式 Markdown 导出：模块 → 功能点 → 用例 → 前置/步骤/预期。

    用例标题前用 `【冒烟】【类型】【优先级】` 标签标注，前置条件、操作步骤、
    预期结果的具体内容都作为子节点展开。兼容 XMind、幕布、飞书大纲等脑图工具的
    Markdown 大纲导入。

    Args:
        title (str): 导出文档的标题。
        rows (list[dict]): build_export_row 生成的导出行列表。
        include_review (bool, optional): 保留参数以兼容旧调用。默认为 False。

    Returns:
        str: 大纲式 Markdown 文本。
    """
    module_order, feature_order, grouped = _group_rows(rows)

    lines = [f"- {title or '测试用例'}"]

    for mod in module_order:
        lines.append(f"  - 模块：{mod}")
        for feat in feature_order[mod]:
            lines.append(f"    - 功能点：{feat}")
            for row in grouped[mod][feat]:
                tags: list[str] = []
                if row.get("is_smoke") == "是":
                    tags.append("冒烟")
                case_type = row.get("case_type", "").strip()
                if case_type:
                    tags.append(case_type)
                priority = row.get("priority", "").strip()
                if priority:
                    tags.append(priority)
                tag_str = "".join(f"【{t}】" for t in tags)
                title_text = (row.get("title", "") or "").strip() or "（未命名用例）"
                lines.append(f"      - {tag_str}{title_text}")

                precondition_lines = _split_text_lines(row.get("precondition", ""))
                if precondition_lines:
                    lines.append("        - 前置条件：")
                    for item in precondition_lines:
                        lines.append(f"          - {item}")

                steps = _split_text_lines(row.get("steps_text", ""))
                if steps:
                    lines.append("        - 操作步骤：")
                    for step in steps:
                        lines.append(f"          - {_strip_step_number(step)}")

                expected_lines = _split_text_lines(row.get("expected_result", ""))
                if expected_lines:
                    lines.append("        - 预期结果：")
                    for item in expected_lines:
                        lines.append(f"          - {item}")

    lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------- Excel


def export_testcases_xlsx(title: str, rows: list[dict], include_review: bool) -> bytes:
    """将规范化后的测试用例行导出为 .xlsx 格式的字节数据。

    生成带紫色表头、冻结首行、自动换行的工作表。

    Args:
        title (str): 工作表名称。
        rows (list[dict]): build_export_row 生成的导出行列表。
        include_review (bool): 是否包含评审状态和来源列。

    Returns:
        bytes: .xlsx 文件的字节数据。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = _columns(include_review)

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "TestCases")[:31] or "TestCases"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, (field, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(field, 16)
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        for c, (field, _) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(field, "") or "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------- dispatch


def export_testcases(
    title: str,
    cases: list[Any],
    fmt: str = "xlsx",
    include_review: bool = False,
) -> tuple[bytes, str, str]:
    """测试用例导出的统一入口，根据格式返回字节内容、媒体类型和文件扩展名。

    调用方需提前将 module 和 feature 补入用例对象，因为 ORM 对象上通常不直接带这两个字段。

    Args:
        title (str): 导出文档的标题。
        cases (list[Any]): 用例对象列表（dict / GeneratedCaseDraft / TestCase）。
        fmt (str, optional): 导出格式，支持 "xlsx" 和 "md"。默认为 "xlsx"。
        include_review (bool, optional): 是否包含评审状态和来源列。默认为 False。

    Returns:
        tuple[bytes, str, str]: (字节内容, MIME 类型, 文件扩展名)。
    """
    rows = [build_export_row(case, idx) for idx, case in enumerate(cases)]
    if fmt == "md":
        text = export_testcases_md(title, rows, include_review)
        return text.encode("utf-8"), MEDIA_MD, "md"
    content = export_testcases_xlsx(title, rows, include_review)
    return content, MEDIA_XLSX, "xlsx"
