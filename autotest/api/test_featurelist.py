"""FeatureList 导入导出：xlsx / md 双格式与往返一致性。"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _export(client, project_id, doc_id, fmt):
    return client.get(
        f"/projects/{project_id}/requirements/{doc_id}/featurelist/export",
        params={"format": fmt},
    )


@pytest.mark.smoke
def test_export_featurelist_xlsx(client, project, confirmed_doc):
    resp = _export(client, project["id"], confirmed_doc["id"], "xlsx")
    assert resp.status_code == 200
    assert XLSX_MIME in resp.headers["content-type"]
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 1 + len(confirmed_doc["items"])  # 表头 + 数据行


def test_export_featurelist_md(client, project, confirmed_doc):
    resp = _export(client, project["id"], confirmed_doc["id"], "md")
    assert resp.status_code == 200
    text = resp.content.decode("utf-8")
    assert confirmed_doc["items"][0]["feature"] in text


def test_export_nonexistent_doc_404(client, project):
    resp = _export(client, project["id"], 9999999, "xlsx")
    assert resp.status_code == 404


def test_import_featurelist_xlsx_roundtrip(client, project, confirmed_doc):
    """导出的 xlsx 再导入，应生成结构一致的新文档。"""
    exported = _export(client, project["id"], confirmed_doc["id"], "xlsx")
    resp = client.post(
        f"/projects/{project['id']}/requirements/featurelist/import",
        files={"file": ("清单.xlsx", exported.content, XLSX_MIME)},
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["source_type"] == "featurelist"
    assert body["status"] == "structured"
    assert len(body["items"]) == len(confirmed_doc["items"])
    assert {i["feature"] for i in body["items"]} == {i["feature"] for i in confirmed_doc["items"]}


def test_import_featurelist_md_roundtrip(client, project, confirmed_doc):
    exported = _export(client, project["id"], confirmed_doc["id"], "md")
    resp = client.post(
        f"/projects/{project['id']}/requirements/featurelist/import",
        files={"file": ("清单.md", exported.content, "text/markdown")},
    )
    assert resp.status_code == 201
    assert len(resp.json()["items"]) == len(confirmed_doc["items"])


def test_import_invalid_featurelist_rejected(client, project):
    resp = client.post(
        f"/projects/{project['id']}/requirements/featurelist/import",
        files={"file": ("bad.xlsx", b"not-a-real-xlsx", XLSX_MIME)},
    )
    assert resp.status_code == 400
